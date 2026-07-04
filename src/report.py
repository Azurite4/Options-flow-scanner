"""Generate a dated Excel report into reports/ — dashboard with verdict,
scored anomalies, volatility metrics, captioned charts, methodology page."""

import sqlite3
from pathlib import Path

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

from detector import (load_universe, bucket_baseline, score_day,
                      flags_for_day, composite, LOOKBACK_DAYS, MIN_HISTORY)
from vol_metrics import day_metrics, load_metrics

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "options.db"
REPORTS = ROOT / "reports"
SCORES_CSV = ROOT / "data" / "daily_scores.csv"

NAVY = "1F3864"
HEADER_FILL = PatternFill("solid", start_color=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
TITLE_FONT = Font(bold=True, size=14, name="Arial", color=NAVY)
SECTION_FONT = Font(bold=True, size=12, name="Arial", color=NAVY)
BODY_FONT = Font(name="Arial")
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="808080")
BIG_METRIC = Font(bold=True, size=22, name="Arial", color=NAVY)
SEV_FILLS = {
    "high": PatternFill("solid", start_color="F8CBAD"),
    "med": PatternFill("solid", start_color="FFE699"),
}


def load_day(target_date: str | None = None) -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    if target_date is None:
        target_date = conn.execute(
            "SELECT MAX(snapshot_date) FROM chain_snapshots").fetchone()[0]
    df = pd.read_sql(
        "SELECT * FROM chain_snapshots WHERE snapshot_date = ?",
        conn, params=[target_date], parse_dates=["snapshot_date", "expiry"],
    )
    conn.close()
    if df.empty:
        raise ValueError(f"No data for {target_date} (not a trading day in the DB?)")
    df["open_interest"] = pd.to_numeric(df["open_interest"], errors="coerce")
    df["volume"] = pd.to_numeric(df["volume"], errors="coerce").fillna(0)
    df["dte"] = (df["expiry"] - df["snapshot_date"]).dt.days
    df["moneyness"] = df["strike"] / df["spot_price"]
    df["vol_oi_ratio"] = df["volume"] / df["open_interest"].replace(0, pd.NA)
    return df


def load_scores() -> pd.DataFrame | None:
    if not SCORES_CSV.exists():
        return None
    return pd.read_csv(SCORES_CSV, parse_dates=["date"])


def run_detector(target: str):
    conn = sqlite3.connect(DB_PATH)
    uni = load_universe(conn)
    conn.close()
    uni = uni[uni["snapshot_date"] <= target]
    if uni.empty:
        return None, None
    actual = uni["snapshot_date"].max()
    today = uni[uni["snapshot_date"] == actual]
    history = uni[uni["snapshot_date"] < actual]
    keep = sorted(history["snapshot_date"].unique())[-LOOKBACK_DAYS:]
    history = history[history["snapshot_date"].isin(keep)]
    if history["snapshot_date"].nunique() < MIN_HISTORY:
        return None, None
    scored = score_day(today, bucket_baseline(history))
    return flags_for_day(scored), composite(scored, history)


def title(ws, text: str):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT


def note(ws, row: int, text: str, col: int = 1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = NOTE_FONT
    c.alignment = Alignment(wrap_text=True, vertical="top")


def styled_header(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_table(ws, start_row: int, frame: pd.DataFrame):
    styled_header(ws, start_row, list(frame.columns))
    for r, row in enumerate(dataframe_to_rows(frame, index=False, header=False),
                            start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val).font = BODY_FONT
    return start_row + 1 + len(frame)


def verdict_text(pct: float | None, comp: dict | None) -> str:
    if pct is None or comp is None:
        return "No historical score context available for this date."
    if pct >= 95:
        level = "EXTREME — comparable to major market events (COVID crash, SVB, Jan-2021 mania)"
    elif pct >= 90:
        level = "VERY UNUSUAL — top decile of all days since 2020"
    elif pct >= 75:
        level = "ELEVATED — busier than roughly 3 out of 4 days"
    elif pct >= 40:
        level = "NORMAL — routine options activity"
    else:
        level = "QUIET — below-average unusual activity"
    tilt = "puts" if comp["put_score"] > comp["call_score"] else "calls"
    return (f"{level}. Anomalous activity today tilts toward {tilt} "
            f"(put score {comp['put_score']:.2f} vs call score {comp['call_score']:.2f}).")


def build_report(df: pd.DataFrame) -> Path:
    REPORTS.mkdir(exist_ok=True)
    snap = df["snapshot_date"].iloc[0].date().isoformat()
    out = REPORTS / f"report_{snap}.xlsx"
    spot = float(df["spot_price"].iloc[0])

    flags, comp = run_detector(snap)
    scores = load_scores()
    pct = rank = None
    row_today = None
    if scores is not None and (scores["date"] == pd.Timestamp(snap)).any():
        row_today = scores.loc[scores["date"] == pd.Timestamp(snap)].iloc[0]
        pct = float((scores["score"] <= row_today["score"]).mean() * 100)
        rank = int((scores["score"] > row_today["score"]).sum()) + 1

    wb = Workbook()
    tmp_charts: list[Path] = []

    # ============ Sheet 1: DASHBOARD ============
    ws = wb.active
    ws.title = "Dashboard"
    title(ws, f"SPY Options Anomaly Report — {snap}")
    note(ws, 2, "This report compares today's options activity against contracts of the "
                "same kind over the trailing 60 trading days, and against every day since 2020.")

    ws["A4"] = "Composite anomaly score"
    ws["A4"].font = BODY_FONT
    ws["A5"] = f"{comp['score']:.3f}" if comp else "n/a"
    ws["A5"].font = BIG_METRIC
    ws["C4"] = "Percentile vs all history"
    ws["C4"].font = BODY_FONT
    ws["C5"] = f"{pct:.0f}th" if pct is not None else "n/a"
    ws["C5"].font = BIG_METRIC
    ws["E4"] = "All-time rank"
    ws["E4"].font = BODY_FONT
    ws["E5"] = f"{rank} of {len(scores)}" if rank is not None else "n/a"
    ws["E5"].font = BIG_METRIC

    ws["A7"] = "Verdict"
    ws["A7"].font = Font(bold=True, name="Arial")
    ws["A8"] = verdict_text(pct, comp)
    ws["A8"].font = BODY_FONT
    ws["A8"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A8:G8")
    ws.row_dimensions[8].height = 30

    call_vol = int(df.loc[df.option_type == "C", "volume"].sum())
    put_vol = int(df.loc[df.option_type == "P", "volume"].sum())
    day_stats = pd.DataFrame({
        "Metric": ["Spot price", "Total volume (contracts)", "Call volume", "Put volume",
                   "Put/Call ratio", "Contracts flagged as anomalous", "Expiries listed"],
        "Value": [round(spot, 2), int(df["volume"].sum()), call_vol, put_vol,
                  round(put_vol / call_vol, 3),
                  len(flags) if flags is not None else "n/a",
                  int(df["expiry"].nunique())],
        "How to read it": [
            "SPY closing price for this session.",
            "All contracts traded across every strike and expiry.",
            "Higher call share often = speculation / upside chasing.",
            "Higher put share often = hedging / downside concern.",
            "SPY normally sits near 1.0–1.3 (institutions hedge with puts). Below ~0.8 is unusual.",
            "Contracts trading far above their bucket's normal volume (see Anomalies sheet).",
            "Number of expiration dates available.",
        ],
    })
    write_table(ws, 10, day_stats)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70

    # ---- Volatility section ----
    vm = day_metrics(df)
    vhist = load_metrics()
    r0 = 20
    ws.cell(row=r0, column=1, value="Volatility (from implied vols)").font = SECTION_FONT

    def _pctile(col, val):
        if vhist is None or pd.isna(val):
            return "n/a"
        s = vhist[col].dropna()
        return f"{(s <= val).mean() * 100:.0f}th percentile since 2020"

    vol_stats = pd.DataFrame({
        "Metric": ["ATM implied vol (20-45 DTE)", "Put-call skew (5% OTM wings)"],
        "Value": [f"{vm['atm_iv']:.3f}" if pd.notna(vm["atm_iv"]) else "n/a",
                  f"{vm['skew']:.4f}" if pd.notna(vm["skew"]) else "n/a"],
        "Context": [_pctile("atm_iv", vm["atm_iv"]), _pctile("skew", vm["skew"])],
        "How to read it": [
            "The market's priced-in expectation of SPY movement — our homemade VIX. "
            "High percentile = market braced for turbulence.",
            "How much more expensive crash insurance (puts) is than upside bets (calls). "
            "Steep skew = fear being priced in, sometimes before volume shows it.",
        ],
    })
    write_table(ws, r0 + 1, vol_stats)
    ws.column_dimensions["D"].width = 70

    # ============ Sheet 2: ANOMALIES ============
    ws2 = wb.create_sheet("Anomalies")
    title(ws2, "Flagged contracts — what made today unusual")
    note(ws2, 2, "Each row is a contract trading far above the normal level for its kind "
                 "(same side, similar % from spot, similar time to expiry). 'Baseline' is the "
                 "median volume for that kind of contract over the last 60 trading days; "
                 "'Excess' is contracts above that. Rows shaded by severity (orange = extreme, "
                 "yellow = strong). 'IV vs normal' above ~1.15 alongside a volume spike suggests "
                 "aggressive buying pressure; near or below 1.0 suggests selling or spread "
                 "activity. OTM/near-ATM contracts only.")

    if flags is None:
        note(ws2, 4, "Not enough trailing history to score this date.")
    elif flags.empty:
        note(ws2, 4, "No anomalies — a quiet, normal day in the speculation zone.")
    else:
        show = flags.head(40)[["option_type", "strike", "expiry", "dte", "dist",
                               "volume", "med", "excess_vol", "robust_z",
                               "implied_vol", "iv_vs_normal"]].copy()
        show["expiry"] = show["expiry"].dt.date.astype(str)
        show["dist"] = (show["dist"] * 100).round(1)
        show[["med", "excess_vol", "robust_z"]] = show[["med", "excess_vol", "robust_z"]].round(1)
        show["implied_vol"] = show["implied_vol"].round(3)
        show["iv_vs_normal"] = pd.to_numeric(show["iv_vs_normal"], errors="coerce").round(2)
        show.columns = ["Type", "Strike", "Expiry", "DTE", "% from spot", "Volume",
                        "Baseline", "Excess", "Robust z", "IV", "IV vs normal"]
        end = write_table(ws2, 4, show)
        for r in range(5, end):
            z = ws2.cell(row=r, column=9).value
            fill = SEV_FILLS["high"] if z and z >= 30 else (
                SEV_FILLS["med"] if z and z >= 10 else None)
            if fill:
                for c in range(1, 12):
                    ws2.cell(row=r, column=c).fill = fill
        ws2.freeze_panes = "A5"
        for col, w in zip("ABCDEFGHIJK", [7, 9, 12, 6, 12, 11, 11, 11, 10, 8, 13]):
            ws2.column_dimensions[col].width = w

    # ============ Sheet 3: CHARTS ============
    ws3 = wb.create_sheet("Charts")
    title(ws3, "Visual overview")

    if scores is not None and row_today is not None:
        fig, ax = plt.subplots(figsize=(9, 3.2))
        ax.plot(scores["date"], scores["score"], lw=0.7, color="#1F3864")
        ax.axhline(scores["score"].quantile(0.90), ls="--", lw=1, color="#C62828",
                   label="90th percentile")
        ax.scatter([row_today["date"]], [row_today["score"]], s=70, zorder=5,
                   color="#C62828", label=f"This report ({snap})")
        ax.set_title("Daily composite anomaly score since 2020")
        ax.legend()
        fig.tight_layout()
        p = REPORTS / "_c_hist.png"
        fig.savefig(p, dpi=110)
        plt.close(fig)
        tmp_charts.append(p)
        ws3.add_image(XLImage(str(p)), "A3")
        note(ws3, 21, "The red dot is this report's day. Days above the dashed line are in the "
                      "top 10% of unusual activity ever recorded by this tool. The big cluster "
                      "in early 2021 is the meme-stock mania — the most anomalous period in the sample.")

    fig, ax = plt.subplots(figsize=(9, 3.2))
    near = df[(df["moneyness"].between(0.9, 1.1)) & (df["dte"] <= 30)]
    for opt, color, lbl in (("C", "#2E7D32", "Calls"), ("P", "#C62828", "Puts")):
        sub = near[near["option_type"] == opt].groupby("strike")["volume"].sum()
        ax.plot(sub.index, sub.values, label=lbl, color=color)
    ax.axvline(spot, ls="--", color="gray", label="Spot")
    ax.set_title("Where the volume sat (±10% of spot, ≤30 DTE)")
    ax.legend()
    fig.tight_layout()
    p = REPORTS / "_c_strike.png"
    fig.savefig(p, dpi=110)
    plt.close(fig)
    tmp_charts.append(p)
    ws3.add_image(XLImage(str(p)), "A24")
    note(ws3, 42, "Volume concentrating above spot = call-side interest (upside bets/chasing). "
                  "Below spot = put-side interest (hedging/downside bets). A big isolated spike "
                  "at one strike is often a single institutional trade.")

    fig, ax = plt.subplots(figsize=(9, 3.0))
    buckets = pd.cut(df["dte"], bins=[-1, 1, 7, 30, 90, 10_000],
                     labels=["0-1d", "2-7d", "8-30d", "31-90d", ">90d"])
    df.groupby(buckets, observed=True)["volume"].sum().plot.bar(ax=ax, color="#1F3864")
    ax.set_title("Volume by days to expiry")
    ax.set_xlabel("")
    fig.tight_layout()
    p = REPORTS / "_c_dte.png"
    fig.savefig(p, dpi=110)
    plt.close(fig)
    tmp_charts.append(p)
    ws3.add_image(XLImage(str(p)), "A45")
    note(ws3, 62, "SPY volume normally concentrates in the shortest expiries. A shift toward "
                  "longer-dated contracts can mean positioning for events further out.")

    # ============ Sheet 4: CHAIN OVERVIEW ============
    ws4 = wb.create_sheet("Chain Overview")
    title(ws4, "Raw top-volume contracts (unfiltered)")
    note(ws4, 2, "The 25 busiest contracts by raw volume, no statistical filter. Expect the "
                 "at-the-money, shortest-dated strikes to dominate here every day — that's "
                 "normal. The Anomalies sheet is where 'unusual' lives. Vol/OI blank = no "
                 "open interest in the historical data source.")
    cols = ["option_type", "strike", "expiry", "dte", "volume",
            "open_interest", "vol_oi_ratio", "implied_vol"]
    top = df.nlargest(25, "volume")[cols].copy()
    top["expiry"] = top["expiry"].dt.date.astype(str)
    top["vol_oi_ratio"] = pd.to_numeric(top["vol_oi_ratio"], errors="coerce").round(2)
    top["implied_vol"] = top["implied_vol"].round(4)
    write_table(ws4, 4, top)
    ws4.freeze_panes = "A5"
    for col in "ABCDEFGH":
        ws4.column_dimensions[col].width = 14

    # ============ Sheet 5: HOW TO READ THIS ============
    ws5 = wb.create_sheet("How To Read This")
    title(ws5, "Methodology, in plain terms")
    lines = [
        ("What the tool does", "Compares every contract's volume against contracts of the same "
         "kind — call/put, similar distance from spot, similar days to expiry — over the "
         "trailing 60 trading days. Robust statistics (median and MAD) so one whale trade "
         "can't distort the baseline."),
        ("Composite score", "Total excess volume across all flagged-eligible contracts, divided "
         "by a normal day's total volume. Score of 1.0 ≈ an entire extra normal-day's worth "
         "of unusual volume."),
        ("Robust z", "How many robust standard deviations above the baseline a contract traded. "
         "Options volume is fat-tailed, so a z of 5+ is common on busy days; the shading uses "
         "10+ and 30+ as meaningful tiers."),
        ("Volatility metrics", "ATM implied vol is the market's priced-in expectation of "
         "movement (a homemade VIX). Skew is the price gap between 5% OTM puts and calls — "
         "a fear gauge that can move before volume does. Both come with percentile context "
         "against every day since 2020."),
        ("What this does NOT do", "Predict direction. An event study over 2020–2023 found no "
         "evidence that high-score days predict which way SPY moves next. Treat this as a "
         "barometer of unusual activity, not a trading signal."),
        ("Data sources", "OptionsDX end-of-day files (2020–2023, no open interest) plus daily "
         "yfinance collection (2026-07 onward, includes OI). Gap between the two eras — "
         "percentile comparisons across eras deserve mild skepticism."),
    ]
    r = 3
    for heading, body in lines:
        ws5.cell(row=r, column=1, value=heading).font = Font(bold=True, name="Arial")
        c = ws5.cell(row=r + 1, column=1, value=body)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws5.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
        ws5.row_dimensions[r + 1].height = 45
        r += 3
    ws5.column_dimensions["A"].width = 100

    wb.save(out)
    for p in tmp_charts:
        p.unlink(missing_ok=True)
    return out


if __name__ == "__main__":
    print(f"Report written: {build_report(load_day())}")